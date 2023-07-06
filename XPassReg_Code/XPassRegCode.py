"""

    Made by Landon Coley. Contact for questions - lcoley@unc.edu
    
"""


import openpyxl

# Ask user for file names
transactions_file = input("Enter the name of the transaction file (e.g., UNCXPASSFareTransactionsJUNE23.xlsx): ")
registrations_file = input("Enter the name of the registration file (e.g., UNCXPassRegistrations.xlsx): ")

# Read the Excel files
transactions_wb = openpyxl.load_workbook(transactions_file)
registrations_wb = openpyxl.load_workbook(registrations_file)

transactions_sheet = transactions_wb.active
registrations_sheet = registrations_wb.active

# Get the maximum row count for each sheet
transactions_max_row = transactions_sheet.max_row
registrations_max_row = registrations_sheet.max_row

# Create sets to store the numbers and statuses
transactions_numbers = set()
registrations_numbers = set()
denied_numbers = set()

# Extract the numbers from transactions sheet
for row in range(2, transactions_max_row + 1):
    cell_value = transactions_sheet.cell(row=row, column=8).value
    if cell_value is not None and cell_value != "Card Number":
        number = str(cell_value).replace("-", "").strip()
        transactions_numbers.add(number)

# Extract the numbers and statuses from registrations sheet
for row in range(2, registrations_max_row + 1):
    number = registrations_sheet.cell(row=row, column=10).value
    status = registrations_sheet.cell(row=row, column=2).value
    if number is not None and status is not None:
        number = str(number).replace("-", "").strip()
        registrations_numbers.add(number)
        if status == 'Denied':
            denied_numbers.add(number)

# Find denied numbers present in transactions
denied_numbers_in_transactions = transactions_numbers.intersection(denied_numbers)

# Find transaction numbers without a match in registrations
unmatched_numbers = transactions_numbers.difference(registrations_numbers)

# Check if all transaction numbers have at least one match in registrations
all_numbers_matched = transactions_numbers.issubset(registrations_numbers)

# Output denied numbers
if len(denied_numbers_in_transactions) > 0:
    print("Denied Numbers:")
    for number in denied_numbers_in_transactions:
        print(number)
else:
    print("No denied numbers found.")

# Output matching status
if all_numbers_matched:
    print("All transaction numbers have at least one match in registrations.")
else:
    print("Not all transaction numbers have a match in registrations.")

# Output unmatched numbers
if len(unmatched_numbers) > 0:
    print("Numbers without a match in registrations:")
    for number in unmatched_numbers:
        print(number)
else:
    print("All transaction numbers have a match in registrations.")


